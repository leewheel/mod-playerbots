#!/usr/bin/env python3
"""Export Chinese/English bot command alias table to Excel."""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CPP = ROOT.parent / "src/Bot/Cmd/ChatHelper.cpp"
OUT = ROOT / "docs/playerbots_command_aliases.xlsx"

# Category hints for whisper commands (by English trigger)
CATEGORY_MAP = {
    "help": "基础",
    "follow": "基础",
    "stay": "基础",
    "attack": "战斗",
    "flee": "战斗",
    "pull": "战斗",
    "pull back": "战斗",
    "ready": "战斗",
    "disperse": "战斗",
    "warning": "战斗",
    "tank attack": "战斗",
    "max dps": "战斗",
    "ra": "战斗",
    "attackers": "战斗",
    "runaway": "战斗",
    "wait for attack time": "战斗",
    "u": "物品装备",
    "c": "物品装备",
    "e": "物品装备",
    "ue": "物品装备",
    "t": "物品装备",
    "nt": "物品装备",
    "s": "物品装备",
    "b": "物品装备",
    "destroy": "物品装备",
    "open items": "物品装备",
    "unlock items": "物品装备",
    "unlock traded item": "物品装备",
    "qi": "物品装备",
    "inv": "物品装备",
    "items": "物品装备",
    "wts": "物品装备",
    "equip upgrade": "物品装备",
    "autogear": "物品装备",
    "autogear bis": "物品装备",
    "ll": "物品装备",
    "ss": "物品装备",
    "add all loot": "物品装备",
    "r": "任务",
    "q": "任务",
    "accept": "任务",
    "drop": "任务",
    "share": "任务",
    "quests": "任务",
    "spells": "法术天赋",
    "spell": "法术天赋",
    "talents": "法术天赋",
    "cast": "法术天赋",
    "castnc": "法术天赋",
    "buff": "法术天赋",
    "gb": "法术天赋",
    "glyphs": "法术天赋",
    "remove glyph": "法术天赋",
    "glyph equip": "法术天赋",
    "save mana": "法术天赋",
    "invite": "社交公会",
    "leave": "社交公会",
    "who": "社交公会",
    "talk": "社交公会",
    "emote": "社交公会",
    "mail": "社交公会",
    "sendmail": "社交公会",
    "ginvite": "社交公会",
    "guild promote": "社交公会",
    "guild demote": "社交公会",
    "guild remove": "社交公会",
    "guild leave": "社交公会",
    "give leader": "社交公会",
    "hire": "社交公会",
    "go": "位置移动",
    "position": "位置移动",
    "teleport": "位置移动",
    "home": "位置移动",
    "range": "位置移动",
    "los": "位置移动",
    "formation": "位置移动",
    "stance": "位置移动",
    "move from group": "位置移动",
    "lfg": "副本PVP",
    "cdebug": "副本PVP",
    "rti": "副本PVP",
    "pull rti": "副本PVP",
    "pvp stats": "副本PVP",
    "ready check": "副本PVP",
    "enter vehicle": "副本PVP",
    "leave vehicle": "副本PVP",
    "repair": "维护经济",
    "bank": "维护经济",
    "trainer": "维护经济",
    "maintenance": "维护经济",
    "craft": "维护经济",
    "drink": "维护经济",
    "grind": "维护经济",
    "calc": "维护经济",
    "roll": "维护经济",
    "tame": "维护经济",
    "pet": "宠物",
    "pet attack": "宠物",
    "rep": "维护经济",
    "stats": "维护经济",
    "aura": "维护经济",
    "emblems": "维护经济",
    "outfit": "维护经济",
    "rtsc": "维护经济",
    "focus heal": "治疗",
    "revive": "治疗",
    "release": "治疗",
    "co": "策略调试",
    "nc": "策略调试",
    "de": "策略调试",
    "cs": "策略调试",
    "debug": "策略调试",
    "chat": "策略调试",
    "log": "策略调试",
    "reset botAI": "策略调试",
    "cheat": "策略调试",
    "wipe": "策略调试",
    "rpg status": "RPG",
    "rpg do quest": "RPG",
    "summon": "召唤",
    "cancel tree form": "德鲁伊形态",
    "cancel travel form": "德鲁伊形态",
    "cancel bear form": "德鲁伊形态",
    "cancel dire bear form": "德鲁伊形态",
    "cancel cat form": "德鲁伊形态",
    "cancel moonkin form": "德鲁伊形态",
    "cancel aquatic form": "德鲁伊形态",
}


def parse_whisper_aliases() -> list[tuple[str, str, str, str]]:
    text = CPP.read_text(encoding="utf-8")
    block = re.search(
        r"static const std::pair<const char\*, const char\*> aliases\[\] = \{(.*?)\};",
        text,
        re.S,
    )
    if not block:
        raise RuntimeError("aliases block not found in ChatHelper.cpp")

    rows = []
    for m in re.finditer(r'\{"([^"]+)", "([^"]+)"\}', block.group(1)):
        zh, en = m.group(1), m.group(2)
        cat = CATEGORY_MAP.get(en, "其他")
        usage = "密语机器人（Whisper）"
        rows.append((cat, zh, en, usage))
    return rows


def style_header(ws, headers):
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_sheet(ws, headers, rows):
    style_header(ws, headers)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, width in enumerate([14, 22, 28, 36], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def main():
    whisper_rows = parse_whisper_aliases()
    whisper_rows.sort(key=lambda x: (x[0], x[2], x[1]))

    pet_rows = [
        ("宠物", "aggressive", "主动姿态"),
        ("宠物", "defensive", "防御姿态"),
        ("宠物", "passive", "被动姿态"),
        ("宠物", "stance", "查看当前姿态"),
        ("宠物", "attack", "攻击目标"),
        ("宠物", "follow", "跟随"),
        ("宠物", "stay", "停留"),
    ]

    bot_rows = [
        ("列表", "list", "列出已控机器人"),
        ("重载/刷新", "reload", "重载配置（GM）"),
        ("微调", "tweak", "调试微调值"),
        ("自己", "self", "开关自我 bot AI"),
        ("查询", "lookup", "查询可用机器人"),
        ("添加", "add", "添加机器人上线"),
        ("添加账号", "addaccount", "按账号添加"),
        ("初始化", "init", "初始化装备（见 init 参数表）"),
        ("初始化自己", "initself", "GM 初始化自己"),
        ("移除/删除", "remove", "移除/登出机器人"),
        ("添加职业/创建职业", "addclass", "创建职业 bot"),
    ]

    init_rows = [
        ("init=auto", "按主人装备自动"),
        ("init=white / init=common", "白装"),
        ("init=green / init=uncommon", "绿装"),
        ("init=blue / init=rare", "蓝装"),
        ("init=epic / init=purple", "紫装"),
        ("init=legendary / init=yellow", "橙装"),
        ("init=数字", "指定装备评分上限"),
        ("refresh=raid", "重置副本 CD（addclass）"),
        ("levelup / level", "升级并随机化"),
        ("refresh", "刷新 bot"),
        ("random", "随机化"),
        ("quests", "初始化副本任务"),
    ]

    class_rows = [
        ("战士", "warrior"),
        ("圣骑士", "paladin"),
        ("猎人", "hunter"),
        ("盗贼", "rogue"),
        ("牧师", "priest"),
        ("萨满", "shaman"),
        ("法师", "mage"),
        ("术士", "warlock"),
        ("德鲁伊", "druid"),
        ("死骑/死亡骑士", "dk"),
    ]

    note_rows = [
        ("密语命令", "中文与英文等价；多词命令需整段输入，如「焦点治疗」「pull back」"),
        ("物品/任务链接", "可直接密语发送物品、任务、物体链接"),
        ("频道前缀", "#w 密语 / #p 队伍 / #r 团队 / #g 公会"),
        ("执行动作", "d 动作名 或 do 动作名（英文）"),
        ("多命令", "用配置 commandSeparator（常为 ;;）分隔"),
        ("服务器命令", ".playerbots bot <子命令> [参数]"),
        ("策略快捷", "战斗=co，非战斗=nc，死亡=de"),
        ("宠物", "主命令用「宠物」，子参数目前以英文为主"),
        ("性别 addclass", "男/male/0，女/female/1"),
        ("数据来源", "ChatHelper.cpp ResolveChatCommandAlias + PlayerbotMgr NormalizeBotSubCommand"),
    ]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "密语命令对照"
    write_sheet(ws1, ["分类", "中文命令", "英文命令", "用法"], whisper_rows)

    ws2 = wb.create_sheet("宠物子命令")
    write_sheet(ws2, ["主命令", "子参数（输入）", "说明"], pet_rows)

    ws3 = wb.create_sheet("playerbots_bot")
    write_sheet(ws3, ["中文", "英文", "说明"], bot_rows)

    ws4 = wb.create_sheet("init参数")
    write_sheet(ws4, ["参数", "说明"], init_rows)

    ws5 = wb.create_sheet("addclass职业")
    write_sheet(ws5, ["中文", "英文"], class_rows)

    ws6 = wb.create_sheet("使用说明")
    write_sheet(ws6, ["项目", "说明"], note_rows)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Written {OUT} ({len(whisper_rows)} whisper aliases)")


if __name__ == "__main__":
    main()
